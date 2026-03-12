"""
AJPL Calculator – Data Safety Backup Engine
AES-256-CBC encrypted .dat backup with Excel snapshot export/import.
"""

import io, os, csv, json, hashlib, zipfile, struct, tempfile
from datetime import datetime, timezone
from typing import Any
import pytz
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import padding as sym_padding, hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.backends import default_backend

IST = pytz.timezone("Asia/Kolkata")
MAGIC = b"AJPLDAT1"
SALT_LEN = 16
IV_LEN = 16
KDF_ITERATIONS = 250_000
APP_NAME = "AJPL Calculator"
APP_VERSION = "1.0.0"
SCHEMA_VERSION = "1"
BACKUP_FORMAT_VERSION = "1"

# Collections for full backup
BUSINESS_COLLECTIONS = [
    "users", "branches", "purities", "rate_cards", "item_names",
    "salespeople", "settings", "customers", "bills",
    "feedback_questions", "feedbacks", "notifications",
]
# Export-only: included in snapshot for audit trail but NOT restored on import
EXPORT_ONLY_COLLECTIONS = ["backup_audit_logs"]
# Ephemeral – excluded from both export and restore
EPHEMERAL_COLLECTIONS = ["otps", "sessions"]
# Transactional collections affected by replace_current_year_data
TRANSACTIONAL_COLLECTIONS = ["bills", "feedbacks", "notifications"]
MASTER_COLLECTIONS = [c for c in BUSINESS_COLLECTIONS if c not in TRANSACTIONAL_COLLECTIONS]

# --------------- crypto helpers ---------------

def _derive_key(password: str, salt: bytes) -> bytes:
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=KDF_ITERATIONS,
        backend=default_backend(),
    )
    return kdf.derive(password.encode("utf-8"))


def encrypt_bytes(plain: bytes, password: str) -> bytes:
    salt = os.urandom(SALT_LEN)
    iv = os.urandom(IV_LEN)
    key = _derive_key(password, salt)
    padder = sym_padding.PKCS7(128).padder()
    padded = padder.update(plain) + padder.finalize()
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    ct = cipher.encryptor().update(padded) + cipher.encryptor().finalize()
    # fix: use single encryptor
    enc = cipher.encryptor()
    ct = enc.update(padded) + enc.finalize()
    return MAGIC + salt + iv + ct


def decrypt_bytes(data: bytes, password: str) -> bytes:
    if len(data) < len(MAGIC) + SALT_LEN + IV_LEN + 16:
        raise ValueError("File too small or corrupt")
    if data[: len(MAGIC)] != MAGIC:
        raise ValueError("Invalid file header (not AJPLDAT1)")
    offset = len(MAGIC)
    salt = data[offset : offset + SALT_LEN]
    offset += SALT_LEN
    iv = data[offset : offset + IV_LEN]
    offset += IV_LEN
    ct = data[offset:]
    key = _derive_key(password, salt)
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    dec = cipher.decryptor()
    padded = dec.update(ct) + dec.finalize()
    unpadder = sym_padding.PKCS7(128).unpadder()
    try:
        return unpadder.update(padded) + unpadder.finalize()
    except Exception:
        raise ValueError("Wrong password or corrupt data")

# --------------- time helpers ---------------

def now_ist():
    return datetime.now(IST)

def year_start_ist(year: int | None = None):
    y = year or now_ist().year
    return IST.localize(datetime(y, 1, 1, 0, 0, 0))

def ist_label(dt=None):
    dt = dt or now_ist()
    return dt.strftime("%Y-%m-%d_%H-%M-%S_IST")

def backup_filename(year, ist_dt):
    return f"AJPL_BACKUP_{year}_UPTO_{ist_label(ist_dt)}.dat"

def excel_filename(year, ist_dt):
    return f"AJPL_EXCEL_SNAPSHOT_{year}_UPTO_{ist_label(ist_dt)}.xlsx"

# --------------- package builder ---------------

def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


async def build_package(db, user_id: str, username: str) -> tuple[dict, bytes]:
    """Build a zip package from DB. Returns (manifest, zip_bytes)."""
    ts = now_ist()
    year = ts.year
    p_start = year_start_ist(year)
    p_end = ts

    buf = io.BytesIO()
    checksums: dict[str, str] = {}
    collection_counts: dict[str, int] = {}

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # --- dump every business collection + export-only collections ---
        all_export_collections = BUSINESS_COLLECTIONS + EXPORT_ONLY_COLLECTIONS
        for col_name in all_export_collections:
            col = db[col_name]
            docs = await col.find({}).to_list(100_000)
            collection_counts[col_name] = len(docs)

            # JSONL
            jsonl_lines = []
            for d in docs:
                d.pop("_id", None)
                jsonl_lines.append(json.dumps(d, default=str))
            jsonl_bytes = "\n".join(jsonl_lines).encode("utf-8")
            jsonl_path = f"data/{col_name}.jsonl"
            zf.writestr(jsonl_path, jsonl_bytes)
            checksums[jsonl_path] = _sha256(jsonl_bytes)

            # CSV
            if docs:
                all_keys = set()
                for d in docs:
                    all_keys.update(d.keys())
                all_keys.discard("_id")
                keys = sorted(all_keys)
                csv_buf = io.StringIO()
                writer = csv.DictWriter(csv_buf, fieldnames=keys, extrasaction="ignore")
                writer.writeheader()
                for d in docs:
                    writer.writerow({k: json.dumps(v, default=str) if isinstance(v, (list, dict)) else v for k, v in d.items() if k != "_id"})
                csv_bytes = csv_buf.getvalue().encode("utf-8")
            else:
                csv_bytes = b""
            csv_path = f"tables/{col_name}.csv"
            zf.writestr(csv_path, csv_bytes)
            checksums[csv_path] = _sha256(csv_bytes)

        # --- Excel workbook ---
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        for col_name in all_export_collections:
            docs = await db[col_name].find({}).to_list(100_000)
            ws = wb.create_sheet(title=col_name[:31])
            if docs:
                keys = sorted({k for d in docs for k in d.keys()} - {"_id"})
                ws.append(keys)
                for d in docs:
                    ws.append([json.dumps(d.get(k), default=str) if isinstance(d.get(k), (list, dict)) else d.get(k, "") for k in keys])
        xl_buf = io.BytesIO()
        wb.save(xl_buf)
        xl_bytes = xl_buf.getvalue()
        zf.writestr("backup.xlsx", xl_bytes)
        checksums["backup.xlsx"] = _sha256(xl_bytes)

        # --- SCHEMA_VERSION ---
        schema = json.dumps({"schema_version": SCHEMA_VERSION, "app_version": APP_VERSION}).encode()
        zf.writestr("SCHEMA_VERSION.json", schema)
        checksums["SCHEMA_VERSION.json"] = _sha256(schema)

        # --- README ---
        readme = _build_readme()
        zf.writestr("README_RECOVERY.txt", readme)
        checksums["README_RECOVERY.txt"] = _sha256(readme.encode())

        # --- manifest ---
        manifest = {
            "app_name": APP_NAME,
            "app_version": APP_VERSION,
            "schema_version": SCHEMA_VERSION,
            "backup_format_version": BACKUP_FORMAT_VERSION,
            "generated_at_ist": ist_label(ts),
            "period_start_ist": p_start.isoformat(),
            "period_end_ist": p_end.isoformat(),
            "collection_counts": collection_counts,
            "checksums": checksums,
            "exported_by_user_id": user_id,
            "exported_by_username": username,
        }
        manifest_bytes = json.dumps(manifest, indent=2).encode()
        zf.writestr("manifest.json", manifest_bytes)
        checksums["manifest.json"] = _sha256(manifest_bytes)

        # --- SHA256SUMS ---
        sums_lines = "\n".join(f"{v}  {k}" for k, v in sorted(checksums.items()) if k != "SHA256SUMS.txt")
        zf.writestr("SHA256SUMS.txt", sums_lines)

    return manifest, buf.getvalue()


async def build_excel_only(db) -> bytes:
    """Build standalone Excel snapshot."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for col_name in BUSINESS_COLLECTIONS + EXPORT_ONLY_COLLECTIONS:
        docs = await db[col_name].find({}).to_list(100_000)
        ws = wb.create_sheet(title=col_name[:31])
        if docs:
            keys = sorted({k for d in docs for k in d.keys()} - {"_id"})
            ws.append(keys)
            for d in docs:
                ws.append([json.dumps(d.get(k), default=str) if isinstance(d.get(k), (list, dict)) else d.get(k, "") for k in keys])
        else:
            ws.append(["(empty collection)"])
    xl_buf = io.BytesIO()
    wb.save(xl_buf)
    return xl_buf.getvalue()

# --------------- import / validation ---------------

def validate_package(zip_bytes: bytes) -> tuple[dict, zipfile.ZipFile]:
    """Validate zip, return manifest and open ZipFile."""
    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes), "r")
    except Exception:
        raise ValueError("Corrupt zip archive")
    if "manifest.json" not in zf.namelist():
        raise ValueError("manifest.json missing from backup")
    manifest = json.loads(zf.read("manifest.json"))
    if manifest.get("schema_version") != SCHEMA_VERSION:
        raise ValueError(f"Schema mismatch: expected {SCHEMA_VERSION}, got {manifest.get('schema_version')}")

    # Validate checksums (mandatory)
    if "SHA256SUMS.txt" not in zf.namelist():
        raise ValueError("SHA256SUMS.txt missing – backup integrity cannot be verified")
    sums_text = zf.read("SHA256SUMS.txt").decode()
    for line in sums_text.strip().split("\n"):
        if not line.strip():
            continue
        expected_hash, fname = line.strip().split("  ", 1)
        if fname in zf.namelist():
            actual = _sha256(zf.read(fname))
            if actual != expected_hash:
                raise ValueError(f"Checksum mismatch for {fname}")
    return manifest, zf


def _load_jsonl(zf: zipfile.ZipFile, col_name: str) -> list[dict]:
    path = f"data/{col_name}.jsonl"
    if path not in zf.namelist():
        return []
    raw = zf.read(path).decode("utf-8")
    docs = []
    for line in raw.strip().split("\n"):
        if line.strip():
            docs.append(json.loads(line))
    return docs


async def import_preview(db, zip_bytes: bytes, mode: str) -> dict:
    """Dry-run preview of import. Returns per-collection action counts."""
    manifest, zf = validate_package(zip_bytes)
    preview = {}

    # Use the backup manifest's period for replace mode (not import-time)
    m_period_start = manifest.get("period_start_ist", "")
    m_period_end = manifest.get("period_end_ist", "")

    for col_name in BUSINESS_COLLECTIONS:
        docs = _load_jsonl(zf, col_name)
        existing_ids = set()
        async for d in db[col_name].find({}, {"id": 1, "_id": 0}):
            if d.get("id"):
                existing_ids.add(d["id"])

        insert_count = 0
        update_count = 0
        skip_count = 0
        delete_count = 0

        if mode == "replace_current_year_data" and col_name in TRANSACTIONAL_COLLECTIONS:
            # Delete DB rows within the backup's period window
            yr_query = {"created_at": {"$gte": m_period_start, "$lte": m_period_end}}
            delete_count = await db[col_name].count_documents(yr_query)
            # Only backup docs within the period window will be inserted;
            # docs outside the window are upserted (merge)
            for d in docs:
                ca = d.get("created_at", "")
                if m_period_start <= ca <= m_period_end:
                    insert_count += 1
                else:
                    did = d.get("id")
                    if did and did in existing_ids:
                        update_count += 1
                    else:
                        insert_count += 1
        else:
            # merge mode
            for d in docs:
                did = d.get("id")
                if did and did in existing_ids:
                    update_count += 1
                elif did:
                    insert_count += 1
                else:
                    insert_count += 1

        preview[col_name] = {
            "insert": insert_count,
            "update": update_count,
            "skip": skip_count,
            "delete": delete_count,
            "backup_count": len(docs),
        }
    zf.close()
    return {"mode": mode, "manifest": manifest, "collections": preview}


async def import_apply(db, zip_bytes: bytes, mode: str) -> dict:
    """Apply import to the database."""
    manifest, zf = validate_package(zip_bytes)
    results = {}

    # Use the backup manifest's period for replace mode
    m_period_start = manifest.get("period_start_ist", "")
    m_period_end = manifest.get("period_end_ist", "")

    for col_name in BUSINESS_COLLECTIONS:
        docs = _load_jsonl(zf, col_name)
        col = db[col_name]
        inserted = 0
        updated = 0
        deleted = 0

        if mode == "replace_current_year_data" and col_name in TRANSACTIONAL_COLLECTIONS:
            # Delete only rows within the backup's period window
            yr_query = {"created_at": {"$gte": m_period_start, "$lte": m_period_end}}
            del_result = await col.delete_many(yr_query)
            deleted = del_result.deleted_count
            # Split docs: those in-window are bulk-inserted, out-of-window are upserted
            in_window = []
            out_window = []
            for d in docs:
                d.pop("_id", None)
                ca = d.get("created_at", "")
                if m_period_start <= ca <= m_period_end:
                    in_window.append(d)
                else:
                    out_window.append(d)
            if in_window:
                await col.insert_many(in_window)
                inserted += len(in_window)
            # Upsert out-of-window docs (e.g. prior-year data) to avoid duplicates
            for d in out_window:
                did = d.get("id")
                if did:
                    existing = await col.find_one({"id": did})
                    if existing:
                        await col.replace_one({"id": did}, d)
                        updated += 1
                    else:
                        await col.insert_one(d)
                        inserted += 1
                else:
                    await col.insert_one(d)
                    inserted += 1
        else:
            # merge / upsert
            for d in docs:
                d.pop("_id", None)
                did = d.get("id")
                if did:
                    existing = await col.find_one({"id": did})
                    if existing:
                        await col.replace_one({"id": did}, d)
                        updated += 1
                    else:
                        await col.insert_one(d)
                        inserted += 1
                else:
                    await col.insert_one(d)
                    inserted += 1

        results[col_name] = {"inserted": inserted, "updated": updated, "deleted": deleted}
    zf.close()
    return {"mode": mode, "results": results}

# --------------- decode instructions ---------------

def _build_readme() -> str:
    return """====================================================
AJPL Calculator – Disaster Recovery / Decode Instructions
====================================================

BACKUP FILE FORMAT: AJPLDAT1
Container: AJPL_BACKUP_<YEAR>_UPTO_<timestamp>.dat

BINARY LAYOUT:
  Bytes 0-7   : Magic header  "AJPLDAT1"
  Bytes 8-23  : Salt           (16 bytes, random)
  Bytes 24-39 : IV             (16 bytes, random)
  Bytes 40+   : Ciphertext     (AES-256-CBC encrypted ZIP)

ENCRYPTION:
  - Algorithm : AES-256-CBC
  - KDF       : PBKDF2-HMAC-SHA256, 250000 iterations
  - Padding   : PKCS7 (block size 128)
  - Key length: 32 bytes (256 bits)

HOW TO DECRYPT (Python 3):
----------------------------------------------------
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import padding, hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.backends import default_backend

password = input("Enter backup password: ")
with open("AJPL_BACKUP_XXXX.dat", "rb") as f:
    data = f.read()

magic = data[:8]   # Should be b"AJPLDAT1"
salt  = data[8:24]
iv    = data[24:40]
ct    = data[40:]

kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32,
                  salt=salt, iterations=250000,
                  backend=default_backend())
key = kdf.derive(password.encode("utf-8"))

cipher = Cipher(algorithms.AES(key), modes.CBC(iv),
                backend=default_backend())
decryptor = cipher.decryptor()
padded = decryptor.update(ct) + decryptor.finalize()

unpadder = padding.PKCS7(128).unpadder()
plaintext = unpadder.update(padded) + unpadder.finalize()

with open("backup.zip", "wb") as f:
    f.write(plaintext)
----------------------------------------------------

AFTER DECRYPTION you have a standard ZIP archive containing:

  manifest.json          – metadata, checksums, period info
  data/<collection>.jsonl – one JSON object per line per collection
  tables/<collection>.csv – same data in CSV for Excel
  backup.xlsx            – multi-sheet Excel workbook
  SHA256SUMS.txt         – integrity checksums
  SCHEMA_VERSION.json    – schema compatibility info
  README_RECOVERY.txt    – this file

OPENING DATA IN EXCEL:
  1. Extract the ZIP
  2. Open backup.xlsx directly, OR
  3. Open individual CSV files from tables/ folder

WARNING:
  The encryption password is NOT stored anywhere in the file.
  Keep your password in a separate secure location.
  Without the password, recovery is impossible.
====================================================
"""


def get_decode_instructions() -> str:
    return _build_readme()
